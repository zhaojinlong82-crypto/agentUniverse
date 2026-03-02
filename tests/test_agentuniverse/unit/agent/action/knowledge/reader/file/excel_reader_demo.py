#!/usr/bin/env python3
# -*- coding:utf-8 -*-

"""
Excel读取器演示示例

这个示例展示了如何使用AgentUniverse的Excel读取器功能来读取和处理Excel文件。
"""

import os
import tempfile
from pathlib import Path

# 创建示例Excel文件
def create_sample_excel():
    """创建一个示例Excel文件用于演示"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息"
        
        # 添加表头
        headers = ["姓名", "年龄", "部门", "工资"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 添加示例数据
        sample_data = [
            ["张三", 25, "技术部", 8000],
            ["李四", 30, "销售部", 9000],
            ["王五", 28, "人事部", 7500],
            ["赵六", 35, "财务部", 8500]
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 创建第二个工作表
        ws2 = wb.create_sheet("部门统计")
        ws2.cell(row=1, column=1, value="部门")
        ws2.cell(row=1, column=2, value="人数")
        ws2.cell(row=2, column=1, value="技术部")
        ws2.cell(row=2, column=2, value=1)
        ws2.cell(row=3, column=1, value="销售部")
        ws2.cell(row=3, column=2, value=1)
        ws2.cell(row=4, column=1, value="人事部")
        ws2.cell(row=4, column=2, value=1)
        ws2.cell(row=5, column=1, value="财务部")
        ws2.cell(row=5, column=2, value=1)
        
        # 保存文件
        file_path = "sample_employees.xlsx"
        wb.save(file_path)
        print(f"✅ 示例Excel文件已创建: {file_path}")
        return file_path
        
    except ImportError:
        print("❌ 需要安装openpyxl: pip install openpyxl")
        return None

def demo_excel_reader():
    """演示Excel读取器功能"""
    print("🚀 开始演示Excel读取器功能...")
    
    # 创建示例文件
    excel_file = create_sample_excel()
    if not excel_file:
        return
    
    try:
        # 导入必要的模块
        from agentuniverse.agent.action.knowledge.reader.file.xlsx_reader import XlsxReader
        from agentuniverse.agent.action.knowledge.reader.file.file_reader import FileReader
        
        print("\n📖 方法1: 直接使用XlsxReader")
        print("-" * 50)
        
        # 直接使用XlsxReader
        xlsx_reader = XlsxReader()
        documents = xlsx_reader.load_data(file=excel_file)
        
        print(f"📊 读取到 {len(documents)} 个文档:")
        for i, doc in enumerate(documents, 1):
            print(f"\n📄 文档 {i} (工作表: {doc.metadata['sheet_name']}):")
            print(f"   文件: {doc.metadata['file_name']}")
            print(f"   行数: {doc.metadata['max_row']}, 列数: {doc.metadata['max_col']}")
            print(f"   内容预览:")
            # 显示前200个字符
            content_preview = doc.text[:200] + "..." if len(doc.text) > 200 else doc.text
            print(f"   {content_preview}")
        
        print("\n📖 方法2: 使用FileReader (自动识别文件类型)")
        print("-" * 50)
        
        # 使用FileReader自动识别文件类型
        file_reader = FileReader()
        documents2 = file_reader.load_data(file_paths=[Path(excel_file)])
        
        print(f"📊 通过FileReader读取到 {len(documents2)} 个文档:")
        for i, doc in enumerate(documents2, 1):
            print(f"\n📄 文档 {i} (工作表: {doc.metadata['sheet_name']}):")
            print(f"   内容长度: {len(doc.text)} 字符")
            print(f"   内容预览:")
            content_preview = doc.text[:150] + "..." if len(doc.text) > 150 else doc.text
            print(f"   {content_preview}")
        
        print("\n✅ Excel读取器演示完成!")
        
    except ImportError as e:
        print(f"❌ 导入错误: {e}")
        print("请确保已安装所需依赖: pip install openpyxl")
    except Exception as e:
        print(f"❌ 发生错误: {e}")
    finally:
        # 清理临时文件
        if os.path.exists(excel_file):
            os.remove(excel_file)
            print(f"\n🧹 已清理临时文件: {excel_file}")

if __name__ == "__main__":
    demo_excel_reader()
