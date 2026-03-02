# !/usr/bin/env python3
# -*- coding:utf-8 -*-

# @Time    : 2024/12/19 10:00
# @Author  : Assistant
# @FileName: xlsx_reader.py
from pathlib import Path
from typing import Union, List, Optional, Dict

from agentuniverse.agent.action.knowledge.reader.reader import Reader
from agentuniverse.agent.action.knowledge.store.document import Document


class XlsxReader(Reader):
    """Excel (.xlsx) file reader.
    
    Used to read and parse Excel format files, supports multiple sheets and various data types.
    """

    def _load_data(self, file: Union[str, Path], ext_info: Optional[Dict] = None) -> List[Document]:
        """Parse Excel file.

        Args:
            file: Excel file path or file object
            ext_info: Additional metadata information

        Returns:
            List[Document]: List of documents containing Excel content

        Note:
            `openpyxl` is required to read Excel files: `pip install openpyxl`
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl is required to read Excel files: "
                "`pip install openpyxl`"
            )

        if isinstance(file, str):
            file = Path(file)

        # Load the workbook
        workbook = openpyxl.load_workbook(file, data_only=True)
        document_list = []

        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Extract data from the worksheet
            sheet_data = []
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Read all data from the worksheet
            for row in range(1, max_row + 1):
                row_data = []
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value is not None:
                        # Convert cell value to string, handling different data types
                        if isinstance(cell.value, (int, float)):
                            row_data.append(str(cell.value))
                        else:
                            row_data.append(str(cell.value))
                    else:
                        row_data.append("")
                
                # Only add non-empty rows
                if any(cell.strip() for cell in row_data):
                    sheet_data.append(" | ".join(row_data))

            # Create document for this sheet
            if sheet_data:
                sheet_content = "\n".join(sheet_data)
                metadata = {
                    "file_name": file.name,
                    "sheet_name": sheet_name,
                    "max_row": max_row,
                    "max_col": max_col
                }
                if ext_info is not None:
                    metadata.update(ext_info)
                
                document_list.append(Document(text=sheet_content, metadata=metadata))

        return document_list
